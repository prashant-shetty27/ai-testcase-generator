from openpyxl import Workbook, load_workbook
from datetime import datetime
from templates.manual_template import ManualTemplate
from templates.automation_template import AutomationTemplate


def export_to_excel(tests, template_type="manual", output_path=None, cases_only=False):

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    if cases_only:
        ws.append([
            "Testcase ID",
            "Priority",
            "Category",
            "Case",
            "Expected Result"
        ])
    else:
        ws.append([
            "Testcase ID",
            "Priority",
            "Category",
            "Scenario",
            "Steps",
            "Example Data",
            "Expected Result"
        ])

    template = AutomationTemplate if template_type == "automation" else ManualTemplate

    for category, cases in tests.items():

        if category == "automation_candidates":
            continue

        if not isinstance(cases, list):
            continue

        for tc in cases:
            if not isinstance(tc, dict):
                continue

            if cases_only:
                ws.append([
                    tc.get("testcase_id"),
                    tc.get("priority", "Medium"),
                    category,
                    tc.get("scenario") or tc.get("title") or "N/A",
                    tc.get("expected_result", ""),
                ])
            else:
                ws.append(template.map_testcase(category, tc))

    if output_path:
        file_path = output_path
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"generated_testcases_{template_type}_{timestamp}.xlsx"

    wb.save(file_path)
    return file_path


def read_existing_testcases(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    headers = []
    rows = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            headers = list(row)
            continue

        case = {headers[i]: value for i, value in enumerate(row)}
        rows.append(case)

    return rows


def normalize_existing_testcases(raw_cases):
    normalized = []

    for row in raw_cases:
        raw_steps = row.get("Steps")
        if isinstance(raw_steps, str):
            steps = [s.strip() for s in raw_steps.splitlines() if s and s.strip()]
        elif raw_steps is None:
            steps = []
        else:
            steps = [str(raw_steps).strip()] if str(raw_steps).strip() else []

        normalized.append({
            "testcase_id": row.get("Testcase ID"),
            "priority": row.get("Priority") or "Medium",
            "category": row.get("Category") or "existing_tests",
            "scenario": row.get("Scenario") or "",
            "steps": steps,
            "expected_result": row.get("Expected Result") or "",
            "examples": row.get("Example Data") or ""
        })

    return normalized
